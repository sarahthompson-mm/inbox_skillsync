#!/usr/bin/env python3
"""
Assembled People Audit
=======================
Pulls all active people from Assembled and produces a clean Excel report
showing roles, teams, sites, queues and flags anything suspicious.

Required environment variables:
  ASSEMBLED_API_KEY  - Assembled API key (sk_live_...)

Output:
  people_audit.xlsx
"""

import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ASSEMBLED_API_KEY = os.environ.get("ASSEMBLED_API_KEY")
ASSEMBLED_BASE    = "https://api.assembledhq.com/v0"
ASSEMBLED_AUTH    = (ASSEMBLED_API_KEY, "")

GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOW = "FFEB9C"
GREY   = "D9D9D9"
BLUE   = "BDD7EE"
PINK   = "FFB3DA"
WHITE  = "FFFFFF"


def assembled_get(path, params=None):
    r = requests.get(f"{ASSEMBLED_BASE}{path}", auth=ASSEMBLED_AUTH, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def assembled_get_all(path, key):
    results = {}
    offset  = 0
    limit   = 100
    while True:
        data = assembled_get(path, params={"limit": limit, "offset": offset})
        page = data.get(key, {})
        if isinstance(page, dict):
            results.update(page)
        else:
            for item in page:
                results[item["id"]] = item
        total   = data.get("total", 0)
        offset += limit
        if offset >= total:
            break
    return results


def fetch_lookup(path, key):
    print(f"Fetching {key}...")
    data = assembled_get_all(path, key)
    lookup = {v["id"]: v["name"] for v in data.values()}
    print(f"  Found {len(lookup)} {key}")
    return lookup


def fetch_people():
    print("Fetching people (all pages)...")
    all_people = assembled_get_all("/people", "people")
    active = [p for p in all_people.values() if not p.get("deleted")]
    print(f"  Found {len(active)} active people")
    return active


def header_row(sheet, headers, colour=GREY):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font      = Font(bold=True, name="Arial")
        cell.fill      = PatternFill("solid", start_color=colour)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def colour_row(sheet, row_num, colour):
    for cell in sheet[row_num]:
        cell.fill = PatternFill("solid", start_color=colour)


def set_column_widths(sheet, widths):
    for i, w in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w


def apply_font(sheet):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if not cell.font or not cell.font.bold:
                cell.font = Font(name="Arial")


def names(uuids, lookup):
    return ", ".join(lookup.get(uid, f"Unknown ({uid[:8]})") for uid in uuids) if uuids else "—"


def build_flags(person, role_lookup, team_lookup, queue_lookup, site_lookup):
    flags = []
    role_names  = [role_lookup.get(r, "") for r in person.get("roles", [])]
    queue_ids   = person.get("queues", [])
    team_ids    = person.get("teams", [])

    is_manager   = any("manager" in r.lower() for r in role_names)
    is_teamlead  = any("lead" in r.lower() for r in role_names)
    is_admin     = any("admin" in r.lower() for r in role_names)
    is_basic     = any("basic" in r.lower() for r in role_names)

    if not queue_ids and not is_manager and not is_admin:
        flags.append("⚠️ No queues assigned")
    if not team_ids:
        flags.append("⚠️ No team assigned")
    if is_manager and queue_ids:
        flags.append("🔍 Manager with queue assignments — intentional?")
    if is_basic and queue_ids:
        flags.append("🔍 Basic role with queues — should this be Standard?")
    if len(team_ids) > 4:
        flags.append(f"🔍 In {len(team_ids)} teams — possible data bloat")
    if not person.get("site"):
        flags.append("⚠️ No site assigned (London/Budapest?)")

    return " | ".join(flags) if flags else "✅ Looks clean"


def build_report(people, role_lookup, team_lookup, queue_lookup, site_lookup):
    wb = Workbook()

    # ── Categorise ────────────────────────────────────────────────────────────
    managers   = []
    teamleads  = []
    admins     = []
    agents     = []
    other      = []

    for p in people:
        role_names = [role_lookup.get(r, "").lower() for r in p.get("roles", [])]
        if any("admin" in r for r in role_names):
            admins.append(p)
        elif any("manager" in r for r in role_names):
            managers.append(p)
        elif any("lead" in r for r in role_names):
            teamleads.append(p)
        elif any("standard" in r or "basic" in r for r in role_names):
            agents.append(p)
        else:
            other.append(p)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Summary"
    ws.append(["Assembled People Audit"])
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws.append([])

    summary = [
        ["Category",           "Count"],
        ["👔 Managers",         len(managers)],
        ["⭐ Team Leads",       len(teamleads)],
        ["🔧 Admins",           len(admins)],
        ["👤 Agents",           len(agents)],
        ["❓ Other / No role",  len(other)],
        ["📊 Total active",     len(people)],
    ]

    for i, row in enumerate(summary):
        ws.append(row)
        row_num = ws.max_row
        if i == 0:
            for cell in ws[row_num]:
                cell.font = Font(bold=True, name="Arial")
                cell.fill = PatternFill("solid", start_color=GREY)
        elif i == 1:
            colour_row(ws, row_num, PINK)
        elif i in (2, 3):
            colour_row(ws, row_num, BLUE)
        elif i == 4:
            colour_row(ws, row_num, GREEN)
        elif i == 5:
            colour_row(ws, row_num, YELLOW)

    set_column_widths(ws, [30, 10])

    # ── Role summary ──────────────────────────────────────────────────────────
    ws.append([])
    ws.append(["Role breakdown"])
    ws[ws.max_row]["A"].font = Font(bold=True, name="Arial")
    ws.append(["Role name", "Count"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, name="Arial")
        cell.fill = PatternFill("solid", start_color=GREY)

    role_counts = {}
    for p in people:
        for r in p.get("roles", []):
            name = role_lookup.get(r, f"Unknown ({r[:8]})")
            role_counts[name] = role_counts.get(name, 0) + 1
    for role_name, count in sorted(role_counts.items(), key=lambda x: -x[1]):
        ws.append([role_name, count])

    # ── Helper to build a people sheet ───────────────────────────────────────
    def people_sheet(title, colour, people_list):
        wsp = wb.create_sheet(title)
        header_row(wsp, [
            "Name", "Email", "Site", "Role(s)",
            "Teams", "Queues", "Flags"
        ], colour)

        for p in sorted(people_list, key=lambda x: f"{x.get('last_name','')} {x.get('first_name','')}"):
            full_name   = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
            site_name   = site_lookup.get(p.get("site", ""), "—")
            role_names  = names(p.get("roles", []), role_lookup)
            team_names  = names(p.get("teams", []), team_lookup)
            queue_names = names(p.get("queues", []), queue_lookup)
            flags       = build_flags(p, role_lookup, team_lookup, queue_lookup, site_lookup)

            wsp.append([
                full_name,
                p.get("email", ""),
                site_name,
                role_names,
                team_names,
                queue_names,
                flags,
            ])

            # Colour rows by flag status
            row_num = wsp.max_row
            if "⚠️" in flags:
                colour_row(wsp, row_num, RED)
            elif "🔍" in flags:
                colour_row(wsp, row_num, YELLOW)
            else:
                colour_row(wsp, row_num, GREEN)

        set_column_widths(wsp, [25, 35, 12, 25, 50, 50, 50])
        apply_font(wsp)
        return wsp

    people_sheet("👔 Managers",        PINK,   managers)
    people_sheet("⭐ Team Leads",       BLUE,   teamleads)
    people_sheet("🔧 Admins",           BLUE,   admins)
    people_sheet("👤 Agents",           GREEN,  agents)
    if other:
        people_sheet("❓ Other",        YELLOW, other)

    # ── Team population sheet ─────────────────────────────────────────────────
    ws_teams = wb.create_sheet("🏢 Teams")
    header_row(ws_teams, ["Team Name", "Member Count", "Members"], GREY)

    team_members = {}
    for p in people:
        full_name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
        for tid in p.get("teams", []):
            tname = team_lookup.get(tid, f"Unknown ({tid[:8]})")
            team_members.setdefault(tname, []).append(full_name)

    for tname, members in sorted(team_members.items(), key=lambda x: -len(x[1])):
        ws_teams.append([tname, len(members), ", ".join(sorted(members))])
        row_num = ws_teams.max_row
        if len(members) == 1:
            colour_row(ws_teams, row_num, YELLOW)

    set_column_widths(ws_teams, [35, 12, 80])
    apply_font(ws_teams)

    return wb, len(managers), len(teamleads), len(admins), len(agents), len(other)


def main():
    if not ASSEMBLED_API_KEY:
        print("Missing ASSEMBLED_API_KEY environment variable")
        sys.exit(1)

    print("═══════════════════════════════════════")
    print("  Assembled People Audit")
    print("═══════════════════════════════════════")

    role_lookup  = fetch_lookup("/roles",  "roles")
    team_lookup  = fetch_lookup("/teams",  "teams")
    queue_lookup = fetch_lookup("/queues", "queues")
    site_lookup  = fetch_lookup("/sites",  "sites")
    people       = fetch_people()

    print("\nBuilding report...")
    wb, managers, leads, admins, agents, other = build_report(
        people, role_lookup, team_lookup, queue_lookup, site_lookup
    )

    output = "people_audit.xlsx"
    wb.save(output)

    print(f"\n── Audit complete ──────────────────────")
    print(f"  👔 Managers:   {managers}")
    print(f"  ⭐ Team Leads: {leads}")
    print(f"  🔧 Admins:     {admins}")
    print(f"  👤 Agents:     {agents}")
    print(f"  ❓ Other:      {other}")
    print(f"  📁 Saved to:   {output}")
    print(f"────────────────────────────────────────")


if __name__ == "__main__":
    main()
