#!/usr/bin/env python3
"""
HiBob → Assembled Dry Run
==========================
Matches HiBob employees (from hibob_employees.csv) to Assembled people
using the HiBob platform ID stored in Assembled's platforms.hibob field.

Does NOT write anything to Assembled — read only, safe to run anytime.

Required environment variables:
  ASSEMBLED_API_KEY  - Assembled API key (sk_live_...)

Required files:
  hibob_employees.csv  - Export from Snowflake BOB_EMPLOYEES table

Output:
  hibob_dryrun.xlsx
"""

import os
import sys
import csv
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ASSEMBLED_API_KEY = os.environ.get("ASSEMBLED_API_KEY")
ASSEMBLED_BASE    = "https://api.assembledhq.com/v0"
ASSEMBLED_AUTH    = (ASSEMBLED_API_KEY, "")

GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOW = "FFEB9C"
GREY   = "D9D9D9"
BLUE   = "BDD7EE"

# Team name overrides — HiBob name (lowercase) → Assembled name (lowercase)
TEAM_NAME_OVERRIDES = {
    "customer happiness": "customer care",
    "retention":          "customer care",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def assembled_get(path, params=None):
    r = requests.get(f"{ASSEMBLED_BASE}{path}", auth=ASSEMBLED_AUTH, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def assembled_get_all(path, key):
    results = {}
    offset  = 0
    limit   = 100
    while True:
        data   = assembled_get(path, params={"limit": limit, "offset": offset})
        page   = data.get(key, {})
        if isinstance(page, dict):
            results.update(page)
        else:
            for item in page:
                results[item["id"]] = item
        total   = data.get("total", 0)
        offset += limit
        if offset >= total:
            break
    return results


def header_row(sheet, headers, colour=GREY):
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font      = Font(bold=True, name="Arial")
        cell.fill      = PatternFill("solid", start_color=colour)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def colour_row(sheet, row_num, colour):
    for cell in sheet[row_num]:
        cell.fill = PatternFill("solid", start_color=colour)


def set_column_widths(sheet, widths):
    for i, w in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w


def apply_font(sheet):
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if not cell.font or not cell.font.bold:
                cell.font = Font(name="Arial")


# ── Fetch HiBob from CSV ──────────────────────────────────────────────────────

def fetch_hibob_employees(csv_path="hibob_employees.csv"):
    print(f"Reading HiBob employees from {csv_path}...")
    if not os.path.exists(csv_path):
        print(f"  ERROR: {csv_path} not found!")
        print("  Please export from Snowflake and commit to the repo.")
        sys.exit(1)

    employees = []
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Normalise column names to lowercase
            emp = {k.lower(): v for k, v in row.items()}
            if emp.get("employee_id") and emp.get("email"):
                employees.append(emp)

    print(f"  Found {len(employees)} active HiBob employees")
    return employees


# ── Fetch Assembled people ────────────────────────────────────────────────────

def fetch_assembled_people():
    print("Fetching Assembled people (all pages)...")
    all_people = assembled_get_all("/people", "people")
    people = {}
    no_hibob_id = 0
    for person in all_people.values():
        if person.get("deleted"):
            continue
        hibob_id = person.get("platforms", {}).get("hibob", "")
        if hibob_id:
            people[str(hibob_id)] = {
                "id":     person["id"],
                "name":   f"{person.get('first_name','')} {person.get('last_name','')}".strip(),
                "email":  person.get("email", ""),
                "teams":  person.get("teams", []),
                "queues": person.get("queues", []),
            }
        else:
            no_hibob_id += 1
    print(f"  Found {len(people)} Assembled people with HiBob platform ID")
    if no_hibob_id:
        print(f"  ⚠ {no_hibob_id} Assembled people have no HiBob platform ID set")
    return people


def fetch_assembled_teams():
    print("Fetching Assembled teams...")
    data = assembled_get_all("/teams", "teams")
    name_to_id = {v["name"].lower().strip(): v["id"] for v in data.values()}
    id_to_name = {v["id"]: v["name"] for v in data.values()}
    print(f"  Found {len(name_to_id)} Assembled teams")
    return name_to_id, id_to_name


# ── Build report ──────────────────────────────────────────────────────────────

def build_report(employees, assembled_people, team_name_to_id, team_id_to_name):
    wb = Workbook()

    matched       = []
    no_match      = []
    name_mismatch = []
    team_changes  = []

    for emp in employees:
        hibob_id  = str(emp.get("employee_id", "")).strip()
        emp_name  = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
        raw_team  = (emp.get("team") or "").strip()

        # Apply team name overrides
        target_team_lower = TEAM_NAME_OVERRIDES.get(raw_team.lower(), raw_team.lower())
        target_team_id    = team_name_to_id.get(target_team_lower)
        overridden        = raw_team.lower() in TEAM_NAME_OVERRIDES
        team_display      = f"{raw_team} → {target_team_lower.title()}" if overridden else raw_team

        person = assembled_people.get(hibob_id)

        if not person:
            no_match.append({
                "name":      emp_name,
                "email":     emp.get("email", ""),
                "hibob_id":  hibob_id,
                "job_title": emp.get("job_title", ""),
                "team":      raw_team,
                "level":     emp.get("level", ""),
            })
            continue

        # Matched!
        matched.append({
            "name":             emp_name,
            "email":            emp.get("email", ""),
            "hibob_id":         hibob_id,
            "job_title":        emp.get("job_title", ""),
            "hibob_team":       team_display,
            "hibob_level":      emp.get("level", ""),
            "manager_email":    emp.get("manager_email", ""),
            "assembled_teams":  ", ".join(team_id_to_name.get(t, t) for t in person["teams"]) or "—",
            "target_team_id":   target_team_id,
            "current_team_ids": person["teams"],
        })

        # Name mismatch?
        if emp_name.lower().strip() != person["name"].lower().strip():
            name_mismatch.append({
                "hibob_name":     emp_name,
                "assembled_name": person["name"],
                "email":          emp.get("email", ""),
                "hibob_id":       hibob_id,
            })

        # Team change needed?
        if target_team_id and target_team_id not in person["teams"]:
            team_changes.append({
                "name":          emp_name,
                "email":         emp.get("email", ""),
                "hibob_team":    team_display,
                "current_teams": ", ".join(team_id_to_name.get(t, t) for t in person["teams"]) or "—",
                "action":        f"Add to '{team_display}' in Assembled",
            })
        elif not target_team_id and raw_team:
            team_changes.append({
                "name":          emp_name,
                "email":         emp.get("email", ""),
                "hibob_team":    raw_team,
                "current_teams": ", ".join(team_id_to_name.get(t, t) for t in person["teams"]) or "—",
                "action":        f"⚠️ HiBob team '{raw_team}' not found in Assembled — needs creating?",
            })

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Summary"
    ws.append(["HiBob → Assembled Dry Run"])
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws.append([])
    ws.append(["READ ONLY — nothing has been changed in Assembled"])
    ws.cell(row=3, column=1).font = Font(italic=True, color="888888", name="Arial")
    ws.append([])

    summary_data = [
        ["Metric",                                    "Count"],
        ["✅ Matched via HiBob platform ID",          len(matched)],
        ["❌ No match found in Assembled",            len(no_match)],
        ["✏️ Name mismatches",                        len(name_mismatch)],
        ["🔄 Team changes needed",                    len(team_changes)],
        ["📊 Total active HiBob employees",           len(employees)],
        ["📊 Total Assembled people with HiBob ID",   len(assembled_people)],
    ]

    for i, row in enumerate(summary_data):
        ws.append(row)
        row_num = ws.max_row
        if i == 0:
            for cell in ws[row_num]:
                cell.font = Font(bold=True, name="Arial")
                cell.fill = PatternFill("solid", start_color=GREY)
        elif i == 1:
            colour_row(ws, row_num, GREEN)
        elif i in (2, 3, 4):
            colour_row(ws, row_num, YELLOW)

    set_column_widths(ws, [45, 10])

    # ── Sheet 2: Matched ──────────────────────────────────────────────────────
    ws_m = wb.create_sheet("✅ Matched")
    header_row(ws_m, [
        "Name", "Email", "HiBob ID", "Job Title",
        "HiBob Team", "Level", "Manager", "Current Assembled Teams"
    ], GREEN)
    for r in sorted(matched, key=lambda x: x["name"]):
        ws_m.append([
            r["name"], r["email"], r["hibob_id"], r["job_title"],
            r["hibob_team"], r["hibob_level"], r["manager_email"], r["assembled_teams"]
        ])
    set_column_widths(ws_m, [25, 35, 25, 35, 25, 8, 35, 50])
    apply_font(ws_m)

    # ── Sheet 3: Team changes needed ──────────────────────────────────────────
    ws_t = wb.create_sheet("🔄 Team Changes")
    header_row(ws_t, ["Name", "Email", "HiBob Team", "Current Assembled Teams", "Action"], YELLOW)
    for r in sorted(team_changes, key=lambda x: x["name"]):
        ws_t.append([r["name"], r["email"], r["hibob_team"], r["current_teams"], r["action"]])
        if "⚠️" in r["action"]:
            colour_row(ws_t, ws_t.max_row, RED)
    set_column_widths(ws_t, [25, 35, 25, 50, 50])
    apply_font(ws_t)

    # ── Sheet 4: Name mismatches ──────────────────────────────────────────────
    ws_n = wb.create_sheet("✏️ Name Mismatches")
    header_row(ws_n, ["HiBob Name", "Assembled Name", "Email", "HiBob ID", "Action"], BLUE)
    for r in sorted(name_mismatch, key=lambda x: x["hibob_name"]):
        ws_n.append([
            r["hibob_name"], r["assembled_name"], r["email"],
            r["hibob_id"], "Check which name is correct"
        ])
    set_column_widths(ws_n, [25, 25, 35, 25, 30])
    apply_font(ws_n)

    # ── Sheet 5: No match ─────────────────────────────────────────────────────
    ws_x = wb.create_sheet("❌ No Match")
    header_row(ws_x, ["Name", "Email", "HiBob ID", "Job Title", "Team", "Level", "Action"], RED)
    for r in sorted(no_match, key=lambda x: x["name"]):
        ws_x.append([
            r["name"], r["email"], r["hibob_id"],
            r["job_title"], r["team"], r["level"],
            "Set HiBob platform ID in Assembled"
        ])
    set_column_widths(ws_x, [25, 35, 25, 35, 25, 8, 38])
    apply_font(ws_x)

    return wb, len(matched), len(no_match), len(name_mismatch), len(team_changes)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not ASSEMBLED_API_KEY:
        print("Missing ASSEMBLED_API_KEY environment variable")
        sys.exit(1)

    print("═══════════════════════════════════════════")
    print("  HiBob → Assembled Dry Run (READ ONLY)")
    print("═══════════════════════════════════════════")

    employees        = fetch_hibob_employees()
    assembled_people = fetch_assembled_people()
    team_name_to_id, team_id_to_name = fetch_assembled_teams()

    print("\nMatching and building report...")
    wb, matched, no_match, name_mismatch, team_changes = build_report(
        employees, assembled_people, team_name_to_id, team_id_to_name
    )

    output = "hibob_dryrun.xlsx"
    wb.save(output)

    print(f"\n── Dry run complete ────────────────────────")
    print(f"  ✅ Matched:             {matched}")
    print(f"  ❌ No match:            {no_match}")
    print(f"  ✏️  Name mismatches:     {name_mismatch}")
    print(f"  🔄 Team changes needed: {team_changes}")
    print(f"  📁 Report saved to:     {output}")
    print(f"  ⚠️  Nothing was changed in Assembled!")
    print(f"────────────────────────────────────────────")


if __name__ == "__main__":
    main()
