#!/usr/bin/env python3
"""
Intercom ↔ Assembled Agent Audit
==================================
Compares agents between Intercom and Assembled and outputs an Excel report
showing who matched, who's missing, and fuzzy suggested fixes by name.

Required environment variables:
  INTERCOM_TOKEN     - Intercom API Bearer token
  ASSEMBLED_API_KEY  - Assembled API key (sk_live_...)

Output:
  agent_audit.xlsx  - Excel report with 5 sheets
"""

import os
import sys
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from thefuzz import process as fuzz_process

INTERCOM_TOKEN    = os.environ.get("INTERCOM_TOKEN")
ASSEMBLED_API_KEY = os.environ.get("ASSEMBLED_API_KEY")

INTERCOM_BASE  = "https://api.intercom.io"
ASSEMBLED_BASE = "https://api.assembledhq.com/v0"

INTERCOM_HEADERS = {
    "Authorization": f"Bearer {INTERCOM_TOKEN}",
    "Intercom-Version": "2.11",
    "Accept": "application/json",
}
ASSEMBLED_AUTH = (ASSEMBLED_API_KEY, "")

GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOW = "FFEB9C"
GREY   = "D9D9D9"
BLUE   = "BDD7EE"


def intercom_get(path):
    r = requests.get(f"{INTERCOM_BASE}{path}", headers=INTERCOM_HEADERS, timeout=30)
    r.raise_for_status()
    return r.json()


def assembled_get(path):
    r = requests.get(f"{ASSEMBLED_BASE}{path}", auth=ASSEMBLED_AUTH, timeout=30)
    r.raise_for_status()
    return r.json()


def fetch_intercom_admins():
    print("Fetching Intercom admins...")
    data = intercom_get("/admins")
    admins = {}
    for admin in data.get("admins", []):
        email = admin.get("email", "").lower().strip()
        if email:
            admins[email] = {
                "name":     admin.get("name", ""),
                "email":    email,
                "team_ids": [str(t) for t in admin.get("team_ids", [])],
            }
    print(f"  Found {len(admins)} Intercom admins")
    return admins


def fetch_intercom_teams():
    print("Fetching Intercom teams...")
    data = intercom_get("/teams")
    teams = {}
    for team in data.get("teams", []):
        teams[str(team["id"])] = team["name"]
    print(f"  Found {len(teams)} Intercom teams")
    return teams


def fetch_assembled_people():
    print("Fetching Assembled people...")
    data = assembled_get("/people")
    people = {}
    for person in data.get("people", {}).values():
        email = person.get("email", "").lower().strip()
        if email:
            people[email] = {
                "name":   f"{person.get('first_name','')} {person.get('last_name','')}".strip(),
                "email":  email,
                "queues": person.get("queues", []),
            }
    print(f"  Found {len(people)} Assembled people")
    return people


def fetch_assembled_queues():
    print("Fetching Assembled queues...")
    data = assembled_get("/queues")
    queues = {}
    for queue in data.get("queues", {}).values():
        queues[queue["id"]] = queue["name"]
    print(f"  Found {len(queues)} Assembled queues")
    return queues


def header_row(sheet, headers, colour=GREY):
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font      = Font(bold=True, name="Arial")
        cell.fill      = PatternFill("solid", start_color=colour)
        cell.alignment = Alignment(horizontal="center")


def colour_row(sheet, row_num, colour):
    for cell in sheet[row_num]:
        cell.fill = PatternFill("solid", start_color=colour)


def set_column_widths(sheet, widths):
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width


def build_report(intercom_admins, intercom_teams, assembled_people, assembled_queues):
    wb = Workbook()

    bot_keywords = ["intercom.io", "facebookbot", "gmail.com", "operator+"]

    matched        = []
    ic_only        = []
    ic_bots        = []
    assembled_only = []

    for email, admin in intercom_admins.items():
        is_bot = any(k in email for k in bot_keywords) or "@" not in email
        if is_bot:
            ic_bots.append(admin)
        elif email in assembled_people:
            matched.append({**admin, **{"assembled": assembled_people[email]}})
        else:
            ic_only.append(admin)

    for email, person in assembled_people.items():
        if email not in intercom_admins:
            assembled_only.append(person)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Intercom ↔ Assembled Agent Audit"])
    ws_summary["A1"].font = Font(bold=True, size=14, name="Arial")
    ws_summary.append([])

    summary_data = [
        ["Metric",                                 "Count"],
        ["✅ Matched (in both systems)",            len(matched)],
        ["⚠️ In Intercom only (real agents)",       len(ic_only)],
        ["🤖 Bots / system accounts (Intercom)",    len(ic_bots)],
        ["⚠️ In Assembled only (not in Intercom)",  len(assembled_only)],
        ["📊 Total Intercom admins",                len(intercom_admins)],
        ["📊 Total Assembled people",               len(assembled_people)],
    ]

    for i, row in enumerate(summary_data):
        ws_summary.append(row)
        if i == 0:
            for cell in ws_summary[ws_summary.max_row]:
                cell.font = Font(bold=True, name="Arial")
                cell.fill = PatternFill("solid", start_color=GREY)
        elif i == 1:
            colour_row(ws_summary, ws_summary.max_row, GREEN)
        elif i in (2, 4):
            colour_row(ws_summary, ws_summary.max_row, YELLOW)

    set_column_widths(ws_summary, [45, 10])

    # ── Sheet 2: Fuzzy name matches ───────────────────────────────────────────
    ws_fuzz = wb.create_sheet("🔍 Suggested Matches")
    header_row(ws_fuzz, [
        "Intercom Name", "Intercom Email",
        "Suggested Assembled Match", "Suggested Email",
        "Confidence", "Intercom Teams", "Action"
    ], BLUE)

    # Build a name → email lookup for Assembled
    assembled_names = {person["name"]: email for email, person in assembled_people.items()}

    fuzzy_rows = []
    for admin in ic_only:
        if not admin["name"]:
            continue
        result = fuzz_process.extractOne(admin["name"], list(assembled_names.keys()))
        if result:
            suggested_name, score = result[0], result[1]
            suggested_email = assembled_names[suggested_name]
            team_names = ", ".join(intercom_teams.get(tid, tid) for tid in admin["team_ids"]) or "No team"
            fuzzy_rows.append((score, admin, suggested_name, suggested_email, team_names))

    high_conf   = 0
    medium_conf = 0

    for score, admin, suggested_name, suggested_email, team_names in sorted(fuzzy_rows, key=lambda x: -x[0]):
        if score >= 80:
            action     = "✅ Likely same person — update email to match"
            row_colour = GREEN
            high_conf += 1
        elif score >= 60:
            action      = "⚠️ Possible match — check manually"
            row_colour  = YELLOW
            medium_conf += 1
        else:
            action     = "❌ Unlikely match — probably needs adding"
            row_colour = RED

        ws_fuzz.append([
            admin["name"],
            admin["email"],
            suggested_name,
            suggested_email,
            f"{score}%",
            team_names,
            action,
        ])
        colour_row(ws_fuzz, ws_fuzz.max_row, row_colour)

    set_column_widths(ws_fuzz, [25, 35, 25, 35, 12, 40, 38])

    # ── Sheet 3: Intercom Only ────────────────────────────────────────────────
    ws_ic = wb.create_sheet("⚠️ Intercom Only")
    header_row(ws_ic, ["Name", "Intercom Email", "Intercom Teams", "Suggested Action"], YELLOW)

    for admin in sorted(ic_only, key=lambda x: x["name"]):
        team_names = ", ".join(intercom_teams.get(tid, tid) for tid in admin["team_ids"]) or "No team"
        ws_ic.append([
            admin["name"],
            admin["email"],
            team_names,
            "See 🔍 Suggested Matches sheet",
        ])

    set_column_widths(ws_ic, [25, 35, 40, 30])

    # ── Sheet 4: Matched ──────────────────────────────────────────────────────
    ws_matched = wb.create_sheet("✅ Matched")
    header_row(ws_matched, ["Name", "Email", "Intercom Teams", "Assembled Queues"], GREEN)

    for m in sorted(matched, key=lambda x: x["name"]):
        team_names  = ", ".join(intercom_teams.get(tid, tid) for tid in m["team_ids"]) or "No team"
        queue_names = ", ".join(assembled_queues.get(qid, qid) for qid in m["assembled"]["queues"]) or "No queue"
        ws_matched.append([m["name"], m["email"], team_names, queue_names])

    set_column_widths(ws_matched, [25, 35, 40, 40])

    # ── Sheet 5: Assembled Only ───────────────────────────────────────────────
    ws_as = wb.create_sheet("⚠️ Assembled Only")
    header_row(ws_as, ["Name", "Assembled Email", "Assembled Queues", "Suggested Action"], YELLOW)

    for person in sorted(assembled_only, key=lambda x: x["name"]):
        queue_names = ", ".join(assembled_queues.get(qid, qid) for qid in person["queues"]) or "No queue"
        ws_as.append([
            person["name"],
            person["email"],
            queue_names,
            "Add to Intercom OR fix email typo",
        ])

    set_column_widths(ws_as, [25, 35, 40, 35])

    for ws in [ws_fuzz, ws_ic, ws_matched, ws_as]:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if not cell.font or not cell.font.bold:
                    cell.font = Font(name="Arial")

    return wb, len(matched), len(ic_only), len(assembled_only), high_conf, medium_conf


def main():
    missing = [v for v in ("INTERCOM_TOKEN", "ASSEMBLED_API_KEY") if not os.environ.get(v)]
    if missing:
        print(f"Missing environment variables: {', '.join(missing)}")
        sys.exit(1)

    print("═══════════════════════════════════════")
    print("  Intercom ↔ Assembled Agent Audit")
    print("═══════════════════════════════════════")

    intercom_admins  = fetch_intercom_admins()
    intercom_teams   = fetch_intercom_teams()
    assembled_people = fetch_assembled_people()
    assembled_queues = fetch_assembled_queues()

    print("\nBuilding report (fuzzy matching on names, bear with...)...")
    wb, matched, ic_only, as_only, high_conf, medium_conf = build_report(
        intercom_admins, intercom_teams, assembled_people, assembled_queues
    )

    output = "agent_audit.xlsx"
    wb.save(output)

    print(f"\n── Audit complete ──────────────────────")
    print(f"  ✅ Matched:                  {matched}")
    print(f"  ⚠️  Intercom only:            {ic_only}")
    print(f"  ⚠️  Assembled only:           {as_only}")
    print(f"  🔍 High confidence matches:  {high_conf} (80%+)")
    print(f"  🔍 Medium confidence:        {medium_conf} (60-79%)")
    print(f"  📁 Report saved to:          {output}")
    print(f"────────────────────────────────────────")


if __name__ == "__main__":
    main()
